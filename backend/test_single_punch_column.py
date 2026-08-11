# -*- coding: utf-8 -*-
import pytest
import pandas as pd
from datetime import date
from openpyxl import load_workbook
from services.attendance_processor import AttendanceProcessor
from services.exporter import AttendanceExporter


def test_single_punch_column_population():
    """
    Test Round 14 Requirement:
    - `SINGLE PUNCH` is populated with actual captured time for all 'Needs Manual Review' rows.
    - `SINGLE PUNCH` is '--' for all other status rows (Present, Absent, Missing Logout, etc.).
    - `First Check In` and `Last Check Out` are None / '--' for 'Needs Manual Review' rows.
    """
    processor = AttendanceProcessor()

    # Employee A0003 test dataset from Round 14 problem statement
    df_raw = pd.DataFrame({
        "Employee ID": ["A0003"] * 5,
        "Employee Name": ["Sumithra"] * 5,
        "Gender": ["Female"] * 5,
        "Department": ["General"] * 5,
        "Attendance Date": ["2026-07-20", "2026-07-21", "2026-07-22", "2026-07-23", "2026-07-24"],
        "First Check In": ["13:54", "13:54", "21:56", "21:59", None],
        "Last Check Out": ["22:15", None, "06:15", "06:11", "06:18"]
    })

    df_punches = pd.DataFrame({
        "Emp ID": ["A0003"] * 5,
        "Date": ["2026-07-20", "2026-07-21", "2026-07-22", "2026-07-23", "2026-07-24"],
        "Time": [
            "13:54, 22:15",
            "13:54",
            "06:15, 21:56",
            "00:19, 06:11, 21:59",
            "06:18"
        ]
    })

    records, cols = processor.process_dataframes(df_raw=df_raw, df_punches=df_punches)
    records.sort(key=lambda r: str(r["attendance_date"]))

    # Verify 'SINGLE PUNCH' is in output column list `cols`
    assert "SINGLE PUNCH" in cols, f"SINGLE PUNCH missing from processor output columns: {cols}"

    for r in records:
        st = r.get("status")
        sp = r.get("SINGLE PUNCH")
        fci = r.get("first_check_in")
        lco = r.get("last_check_out")

        if st == "Needs Manual Review":
            assert sp not in (None, "", "--"), f"Row status is {st} but SINGLE PUNCH is {sp}"
            assert fci is None or fci == "--", f"Needs Manual Review row should have blank First Check In, got {fci}"
            assert lco is None or lco == "--", f"Needs Manual Review row should have blank Last Check Out, got {lco}"
        else:
            assert sp in (None, "", "--"), f"Non-manual-review row status is {st} but SINGLE PUNCH is {sp}"

    # Specific check for A0003 on 2026-07-21 (Example from Round 14 prompt)
    r_21 = next(r for r in records if str(r["attendance_date"]) == "2026-07-21")
    assert r_21["status"] == "Needs Manual Review"
    assert r_21["SINGLE PUNCH"] == "13:54"
    assert r_21["single_punch"] == "13:54"
    assert r_21["first_check_in"] is None
    assert r_21["last_check_out"] is None


def test_exporter_single_punch_column():
    """
    Test that AttendanceExporter.export_to_excel includes SINGLE PUNCH column
    positioned near First Check In / Last Check Out.
    """
    processor = AttendanceProcessor()
    df_raw = pd.DataFrame({
        "Employee ID": ["A0003", "A0001"],
        "Employee Name": ["Sumithra", "John"],
        "Attendance Date": ["2026-07-21", "2026-07-21"],
        "First Check In": ["13:54", "09:00"],
        "Last Check Out": [None, "17:30"]
    })

    records, cols = processor.process_dataframes(df_raw=df_raw)
    excel_buf = AttendanceExporter.export_to_excel(records, columns=cols)

    wb = load_workbook(excel_buf)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    assert "SINGLE PUNCH" in headers, f"Excel headers missing SINGLE PUNCH: {headers}"

    sp_col_idx = headers.index("SINGLE PUNCH") + 1

    # Find A0003 row (Needs Manual Review)
    row_a3 = None
    row_a1 = None
    for row in range(2, ws.max_row + 1):
        emp_id = ws.cell(row=row, column=1).value
        if emp_id == "A0003":
            row_a3 = row
        elif emp_id == "A0001":
            row_a1 = row

    assert row_a3 is not None, "A0003 row not found in exported Excel"
    assert ws.cell(row=row_a3, column=sp_col_idx).value == "13:54"

    if row_a1:
        assert ws.cell(row=row_a1, column=sp_col_idx).value == "--"
