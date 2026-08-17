import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import date, datetime
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors


class AttendanceExporter:

    # ── shared style constants ────────────────────────────────────────────────
    _HDR_FILL   = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")  # Dark navy
    _HDR_FONT   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    _SUM_FILL   = PatternFill(start_color="14532D", end_color="14532D", fill_type="solid")  # Dark emerald
    _SUM_FONT   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    _NRM_FONT   = Font(name="Calibri", size=10, color="1E293B")   # Normal data font
    _BOLD_FONT  = Font(name="Calibri", size=10, bold=True, color="1E293B")
    _CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=False)
    _LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=False, indent=1)
    _LEFT_WRAP  = Alignment(horizontal="left",   vertical="top",    wrap_text=True,  indent=1)
    _THIN_BORD  = Border(
        left=Side(style="thin", color="C0C8D4"),
        right=Side(style="thin", color="C0C8D4"),
        top=Side(style="thin", color="C0C8D4"),
        bottom=Side(style="thin", color="C0C8D4"),
    )
    _MED_BORD   = Border(
        left=Side(style="medium", color="94A3B8"),
        right=Side(style="medium", color="94A3B8"),
        top=Side(style="medium", color="94A3B8"),
        bottom=Side(style="medium", color="94A3B8"),
    )

    # ── row-level fills (Daily Detail) ────────────────────────────────────────
    _FILL_GREEN  = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")  # green-100
    _FONT_GREEN  = Font(name="Calibri", size=10, color="14532D", bold=True)  # green-900
    _FILL_PURPLE = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")  # violet-100
    _FONT_PURPLE = Font(name="Calibri", size=10, color="4C1D95", bold=True)  # violet-900
    _FILL_YELLOW = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")  # yellow-100
    _FONT_YELLOW = Font(name="Calibri", size=10, color="713F12", bold=True)  # yellow-900
    _FILL_ORANGE = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")  # orange-100
    _FONT_ORANGE = Font(name="Calibri", size=10, color="7C2D12", bold=True)  # orange-900
    _FILL_RED    = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid")  # rose-100
    _FONT_RED    = Font(name="Calibri", size=10, color="881337", bold=True)  # rose-900

    # ── zebra fills (Monthly Summary) ────────────────────────────────────────
    _ZEBRA_ODD  = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")  # slate-50
    _ZEBRA_EVEN = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # =========================================================================
    # VALUE HELPER
    # =========================================================================

    @classmethod
    def _get_val(cls, rec, key, default=""):
        if not isinstance(rec, dict):
            val = getattr(rec, key, default)
            return val if val is not None else default

        if key in rec and rec[key] is not None and rec[key] != "":
            return rec[key]

        key_lower = str(key).lower().replace("_", "").replace(".", "").replace(" ", "")

        mappings = {
            "no":                  ["NO.", "No.", "NO", "No", "S.No", "Sl.No", "raw_idx"],
            "empid":               ["employee_id", "EMPLOYEE ID", "Emp ID", "EMP ID", "EMP CODE", "Emp Code", "Staff ID", "User ID"],
            "employeeid":          ["employee_id", "EMPLOYEE ID", "Emp ID", "EMP ID", "EMP CODE", "Emp Code", "Staff ID", "User ID"],
            "employeename":        ["employee_name", "EMPLOYEE NAME", "First Name", "FIRST NAME", "Name", "NAME", "Staff Name"],
            "firstname":           ["employee_name", "EMPLOYEE NAME", "First Name", "FIRST NAME", "Name", "NAME"],
            "name":                ["employee_name", "EMPLOYEE NAME", "First Name", "FIRST NAME", "Name", "NAME"],
            "gender":              ["gender", "GENDER", "Gender", "Sex"],
            "department":          ["department", "DEPARTMENT", "Dept", "DEPT"],
            "date":                ["attendance_date", "DATE", "Date", "Attendance Date"],
            "attendancedate":      ["attendance_date", "DATE", "Date"],
            "logoutdate":          ["logout_date_str", "logout_date", "LOGOUT DATE", "Logout Date", "Check-Out Date"],
            "weekday":             ["weekday", "WEEKDAY", "Day", "DAY"],
            "shift":               ["shift", "Shift", "SHIFT", "shifts", "Shifts", "SHIFTS"],
            "firstcheckin":        ["first_check_in", "FIRST CHECK IN", "First Check In", "Check In", "In Time"],
            "checkin":             ["first_check_in", "FIRST CHECK IN", "First Check In", "Check In", "In Time"],
            "lastcheckout":        ["last_check_out", "LAST CHECK OUT", "Last Check Out", "Check Out", "Out Time"],
            "checkout":            ["last_check_out", "LAST CHECK OUT", "Last Check Out", "Check Out", "Out Time"],
            "singlepunch":         ["SINGLE PUNCH", "Single Punch", "single_punch", "SINGLE PUNCH TIME", "Single Punch Time"],
            "workinghours":        ["working_hours", "WORKING HOURS", "Working Hours", "Total Time", "TOTAL TIME"],
            "overtimehours":       ["overtime_hours", "OVERTIME HOURS", "Overtime Hours", "OT Hours", "OT HOURS", "Overtime", "OVERTIME", "OT", "ot_hours"],
            "othours":             ["overtime_hours", "OVERTIME HOURS", "Overtime Hours", "OT Hours", "OT HOURS", "Overtime", "OVERTIME", "OT", "ot_hours"],
            "overtime":            ["overtime_hours", "OVERTIME HOURS", "Overtime Hours", "OT Hours", "OT HOURS", "Overtime", "OVERTIME", "OT", "ot_hours"],
            "status":              ["status", "Status", "STATUS"],
            "punchstatus":         ["PUNCH STATUS", "Punch Status", "punch_status", "MISSING SHIFT DETAILS", "Missing Shift Details"],
            "missingshiftdetails": ["PUNCH STATUS", "Punch Status", "punch_status", "MISSING SHIFT DETAILS", "Missing Shift Details"],
            "remarks":             ["remarks", "Remarks", "REMARKS"],
        }

        if key_lower in mappings:
            for candidate in mappings[key_lower]:
                if candidate in rec and rec[candidate] is not None and rec[candidate] != "":
                    return rec[candidate]

        for k, v in rec.items():
            k_clean = str(k).lower().replace("_", "").replace(".", "").replace(" ", "")
            if k_clean == key_lower and v is not None and v != "":
                return v

        return default

    # =========================================================================
    # WORKING-HOURS HELPERS
    # =========================================================================

    @classmethod
    def _parse_hhmm_to_mins(cls, val):
        """Parse 'HH:MM' string or decimal float to total minutes. Returns 0 on failure."""
        if val is None or str(val).strip() in ("", "--", "None", "nan", "00:00"):
            return 0
        val_s = str(val).strip()
        if ":" in val_s:
            parts = val_s.split(":")
            try:
                return int(parts[0]) * 60 + int(parts[1])
            except Exception:
                return 0
        try:
            f = float(val_s)
            return int(round(f * 60))
        except Exception:
            return 0

    @classmethod
    def _mins_to_hhmm(cls, total_mins):
        """Convert total minutes integer to 'HHH:MM' string."""
        if total_mins <= 0:
            return "00:00"
        h = total_mins // 60
        m = total_mins % 60
        return f"{h:02d}:{m:02d}"

    # =========================================================================
    # MONTHLY SUMMARY BUILDER
    # =========================================================================

    @classmethod
    def _build_monthly_summary(cls, records):
        """
        Aggregate already-processed daily records into one row per employee.
        Returns list of dicts with Monthly Summary columns, sorted by Employee ID.
        Only records belonging to the primary attendance month are aggregated to prevent
        cross-month leakage (e.g. Aug 1 day-shift records inflating July total days to 32).
        Overnight shifts starting in July with logout on Aug 1 remain attendance_date=July 31
        and are fully preserved in the July summary.
        """
        from collections import defaultdict, Counter

        # Extract primary attendance date from each record
        def extract_date(rec):
            d = None
            if isinstance(rec, dict):
                d = rec.get("attendance_date") or rec.get("date") or rec.get("DATE") or rec.get("Date") or rec.get("Attendance Date")
            else:
                d = getattr(rec, "attendance_date", None) or getattr(rec, "date", None)

            if d is None or str(d).strip() in ("", "--", "None", "nan"):
                d = cls._get_val(rec, "attendance_date") or cls._get_val(rec, "date")

            if d is None or str(d).strip() in ("", "--", "None", "nan"):
                return None

            if isinstance(d, date) and not isinstance(d, datetime):
                return d
            if isinstance(d, datetime):
                return d.date()

            d_str = str(d).strip()
            if " " in d_str:
                d_str = d_str.split()[0]

            for fmt in (
                "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d",
                "%d.%m.%Y", "%Y.%m.%d", "%d-%b-%Y", "%d-%B-%Y", "%Y%m%d",
                "%d%m%Y", "%d-%m-%y", "%d/%m/%y", "%d.%m.%y",
            ):
                try:
                    return datetime.strptime(d_str, fmt).date()
                except ValueError:
                    pass

            try:
                import pandas as pd
                dt = pd.to_datetime(d_str, dayfirst=True)
                if not pd.isna(dt):
                    return dt.date()
            except Exception:
                pass

            return None

        # Determine primary target month (year, month) across all records
        all_att_dates = [extract_date(r) for r in records if extract_date(r) is not None]
        if all_att_dates:
            primary_ym = Counter((d.year, d.month) for d in all_att_dates).most_common(1)[0][0]
        else:
            primary_ym = None

        groups = defaultdict(list)
        for rec in records:
            emp_id = (
                cls._get_val(rec, "employee_id")
                or cls._get_val(rec, "Emp ID")
                or cls._get_val(rec, "EMP CODE")
                or "__unknown__"
            )
            groups[str(emp_id).strip()].append(rec)

        # Global date span scoped to primary month
        global_dates = [d for d in all_att_dates if (primary_ym is None or (d.year, d.month) == primary_ym)]
        global_min = min(global_dates) if global_dates else (min(all_att_dates) if all_att_dates else None)
        global_max = max(global_dates) if global_dates else (max(all_att_dates) if all_att_dates else None)
        total_days_global = (global_max - global_min).days + 1 if (global_min and global_max) else 0

        summary_rows = []

        for emp_id in sorted(groups.keys(), key=lambda x: (x[0].upper(), x)):
            rows = groups[emp_id]

            def first_val(field):
                for r in rows:
                    v = cls._get_val(r, field)
                    if v and str(v).strip() not in ("", "--", "None", "nan"):
                        return str(v).strip()
                return "--"

            emp_name = first_val("employee_name") or first_val("First Name") or first_val("FIRST NAME")
            gender   = first_val("gender")

            # Per-employee calendar span scoped to primary month
            emp_dates = []
            scoped_rows = []
            for rec in rows:
                rec_date = extract_date(rec)
                if rec_date is None:
                    continue
                if primary_ym is None or (rec_date.year, rec_date.month) == primary_ym:
                    emp_dates.append(rec_date)
                    scoped_rows.append(rec)

            # If no rows fall in primary month for this employee, fallback to all employee rows
            if not scoped_rows:
                scoped_rows = rows
                emp_dates = [extract_date(r) for r in rows if extract_date(r) is not None]

            emp_total_days = (
                (max(emp_dates) - min(emp_dates)).days + 1 if emp_dates else total_days_global
            )

            # Counters
            present_days  = absent_days  = half_days = 0
            shift_a       = shift_general = shift_b     = shift_b1 = shift_c = 0
            needs_review = 0
            total_wh_mins = total_ot_mins = 0

            for rec in scoped_rows:
                st  = str(cls._get_val(rec, "status") or "").lower().strip()
                sft = str(cls._get_val(rec, "shift")  or "").strip().upper()
                clean_sft = sft.replace("SHIFT", "").replace("(AUTO-DETECTED)", "").replace("AUTO-DETECTED", "").strip()

                is_present = "present" in st or "short hours" in st or "full day" in st

                if is_present:
                    present_days += 1
                    if "half day" in st:
                        half_days += 1
                else:
                    absent_days += 1

                if is_present:
                    if clean_sft.startswith("A") or clean_sft == "1":
                        shift_a += 1
                    elif clean_sft.startswith("GEN") or clean_sft == "4":
                        shift_general += 1
                    elif clean_sft.startswith("B1") or clean_sft == "5":
                        shift_b1 += 1
                    elif clean_sft.startswith("B") or clean_sft == "2":
                        shift_b += 1
                    elif clean_sft.startswith("C") or "NIGHT" in clean_sft or clean_sft == "3":
                        shift_c += 1

                if "manual review" in st or "needs manual" in st or "single punch" in st:
                    needs_review += 1

                # Working hours
                wh_dec = (
                    rec.get("working_hours_decimal")
                    if isinstance(rec, dict)
                    else getattr(rec, "working_hours_decimal", None)
                )
                if wh_dec is not None:
                    try:
                        total_wh_mins += int(round(float(wh_dec) * 60))
                    except Exception:
                        total_wh_mins += cls._parse_hhmm_to_mins(cls._get_val(rec, "working_hours"))
                else:
                    total_wh_mins += cls._parse_hhmm_to_mins(cls._get_val(rec, "working_hours"))

                # Overtime hours
                ot_dec = (
                    rec.get("overtime_hours_decimal")
                    if isinstance(rec, dict)
                    else getattr(rec, "overtime_hours_decimal", None)
                )
                if ot_dec is not None:
                    try:
                        total_ot_mins += int(round(float(ot_dec) * 60))
                    except Exception:
                        total_ot_mins += cls._parse_hhmm_to_mins(cls._get_val(rec, "overtime_hours"))
                else:
                    total_ot_mins += cls._parse_hhmm_to_mins(cls._get_val(rec, "overtime_hours"))

            summary_rows.append({
                "Employee ID":               emp_id,
                "First Name":                emp_name,
                "Gender":                    gender,
                "Total Days":                emp_total_days,
                "Present Days":              present_days,
                "Absent Days":               absent_days,
                "Half Days":                 half_days,
                "Shift A Count":             shift_a,
                "General Count":             shift_general,
                "Shift B Count":             shift_b,
                "Shift B1 Count":            shift_b1,
                "Shift C Count":             shift_c,
                "Needs Manual Review Count": needs_review,
                "Total Working Hours":       cls._mins_to_hhmm(total_wh_mins),
                "Total Overtime Hours":      cls._mins_to_hhmm(total_ot_mins),
            })

        return summary_rows

    # =========================================================================
    # SHARED SHEET WRITER
    # =========================================================================

    @classmethod
    def _write_sheet(cls, ws, headers, rows,
                     header_fill, header_font,
                     status_col_name=None,
                     zebra=False):
        """Write headers + rows to ws with styling. Freezes row 2. Auto-sizes columns."""
        ws.append(headers)
        col_max_lens = [len(str(h)) for h in headers]

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = cls._CENTER

        status_col_idx = None
        if status_col_name:
            for i, h in enumerate(headers, start=1):
                if str(h).lower() == status_col_name.lower():
                    status_col_idx = i
                    break

        for row_idx, rec in enumerate(rows, start=2):
            row_data = []
            for col_i, h in enumerate(headers):
                val = cls._get_val(rec, h)
                val_str = str(val) if val is not None and val != "" else ""
                row_data.append(val_str)
                if len(val_str) > col_max_lens[col_i]:
                    col_max_lens[col_i] = len(val_str)
            ws.append(row_data)

            if zebra:
                fill = cls._ZEBRA_ODD if row_idx % 2 == 0 else cls._ZEBRA_EVEN
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col_num)
                    cell.fill      = fill
                    cell.alignment = cls._CENTER

            if status_col_idx:
                sc = ws.cell(row=row_idx, column=status_col_idx)
                sc.alignment = cls._CENTER
                st = str(cls._get_val(rec, "status")).lower()
                if "present" in st:
                    sc.fill = cls._FILL_GREEN;  sc.font = cls._FONT_GREEN
                elif "manual review" in st or "needs manual review" in st:
                    sc.fill = cls._FILL_PURPLE; sc.font = cls._FONT_PURPLE
                elif "missing" in st:
                    sc.fill = cls._FILL_YELLOW; sc.font = cls._FONT_YELLOW
                elif "overtime" in st:
                    sc.fill = cls._FILL_ORANGE; sc.font = cls._FONT_ORANGE
                else:
                    sc.fill = cls._FILL_RED;    sc.font = cls._FONT_RED

        for col_i, max_len in enumerate(col_max_lens, start=1):
            ws.column_dimensions[get_column_letter(col_i)].width = max(min(max_len + 4, 50), 12)

        ws.freeze_panes = ws["A2"]

    # =========================================================================
    # PUBLIC — EXCEL EXPORT (two sheets)
    # =========================================================================

    @classmethod
    def export_to_excel(cls, records, columns=None):
        """
        Generate a two-sheet Excel workbook:
          Sheet 1 - 'Daily Detail'    : per-row output (unchanged content)
          Sheet 2 - 'Monthly Summary' : one row per employee aggregated from records
        Both sheets reflect whatever records list is passed in, so filtered
        exports produce a filtered summary automatically.
        """
        wb = openpyxl.Workbook()

        # ── Sheet 1: Daily Detail ─────────────────────────────────────────────
        ws_detail = wb.active
        ws_detail.title = "Daily Detail"

        if not records:
            detail_headers = columns or [
                "Employee ID", "Employee Name", "Department", "Date", "Day",
                "Shift", "First Check In", "Last Check Out", "SINGLE PUNCH",
                "Working Hours", "Overtime Hours", "Status",
            ]
        elif columns:
            detail_headers = list(columns)

            has_ot = any(
                str(h).lower().replace("_", "").replace(" ", "") in
                ("overtimehours", "othours", "overtime", "ot")
                for h in detail_headers
            )
            if not has_ot:
                wh_idx = next(
                    (i for i, h in enumerate(detail_headers)
                     if str(h).lower().replace("_", "").replace(" ", "") in
                     ("workinghours", "totaltime", "totalhours")),
                    -1,
                )
                if wh_idx != -1:
                    detail_headers.insert(wh_idx + 1, "Overtime Hours")
                else:
                    detail_headers.append("Overtime Hours")

            has_sp = any(
                str(h).lower().replace("_", "").replace(" ", "") in
                ("singlepunch", "singlepunchtime", "unpairedpunch")
                for h in detail_headers
            )
            if not has_sp:
                lco_idx = next(
                    (i for i, h in enumerate(detail_headers)
                     if str(h).lower().replace("_", "").replace(" ", "") in
                     ("lastcheckout", "checkout", "outtime")),
                    -1,
                )
                if lco_idx != -1:
                    detail_headers.insert(lco_idx + 1, "SINGLE PUNCH")
                else:
                    fci_idx = next(
                        (i for i, h in enumerate(detail_headers)
                         if str(h).lower().replace("_", "").replace(" ", "") in
                         ("firstcheckin", "checkin", "intime")),
                        -1,
                    )
                    if fci_idx != -1:
                        detail_headers.insert(fci_idx + 1, "SINGLE PUNCH")
                    else:
                        detail_headers.append("SINGLE PUNCH")

        else:
            detail_headers = [
                "Employee ID", "Employee Name", "Department", "Date", "Day",
                "Shift", "First Check In", "Last Check Out", "SINGLE PUNCH",
                "Working Hours", "Overtime Hours", "Status",
            ]

        cls._write_sheet(
            ws_detail, detail_headers, records,
            header_fill=cls._HDR_FILL,
            header_font=cls._HDR_FONT,
            status_col_name="status",
            zebra=False,
        )

        # ── Sheet 2: Monthly Summary ──────────────────────────────────────────
        ws_summary = wb.create_sheet("Monthly Summary")

        summary_headers = [
            "Employee ID",
            "First Name",
            "Gender",
            "Total Days",
            "Present Days",
            "Absent Days",
            "Half Days",
            "Shift A Count",
            "General Count",
            "Shift B Count",
            "Shift B1 Count",
            "Shift C Count",
            "Needs Manual Review Count",
            "Total Working Hours",
            "Total Overtime Hours",
        ]

        summary_rows = cls._build_monthly_summary(records) if records else []

        cls._write_sheet(
            ws_summary, summary_headers, summary_rows,
            header_fill=cls._SUM_FILL,
            header_font=cls._SUM_FONT,
            status_col_name=None,
            zebra=True,
        )

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # =========================================================================
    # PUBLIC — PDF EXPORT (unchanged)
    # =========================================================================

    @classmethod
    def export_to_pdf(cls, records):
        """Generates a styled PDF report for processed attendance records."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=landscape(letter),
            rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30,
        )
        story = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=18,
            leading=22,
            textColor=colors.HexColor("#1E293B"),
            spaceAfter=10,
        )
        subtitle_style = ParagraphStyle(
            "ReportSubTitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#64748B"),
            spaceAfter=15,
        )

        story.append(Paragraph("Smart Attendance Summary Report", title_style))
        story.append(Paragraph(
            "Automated shift detection, overnight punch pairing, and anomaly audit report.",
            subtitle_style,
        ))

        table_data = [[
            "Emp ID", "Employee", "Date", "Shift",
            "Check In", "Check Out", "Single Punch",
            "Hours", "OT Hours", "Status", "Remarks",
        ]]

        for r in records:
            table_data.append([
                cls._get_val(r, "employee_id"),
                str(cls._get_val(r, "employee_name"))[:15],
                str(cls._get_val(r, "attendance_date")),
                cls._get_val(r, "shift"),
                cls._get_val(r, "first_check_in", "--"),
                cls._get_val(r, "last_check_out", "--"),
                cls._get_val(r, "single_punch", "--"),
                cls._get_val(r, "working_hours", "00:00"),
                cls._get_val(r, "overtime_hours", "00:00"),
                cls._get_val(r, "status"),
                str(cls._get_val(r, "remarks", ""))[:25],
            ])

        t = Table(table_data, colWidths=[55, 80, 65, 35, 50, 50, 55, 45, 45, 75, 145])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#1E293B")),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",      (0, 1), (-1, -1), 8),
            ("TOPPADDING",    (0, 1), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
        ]))

        story.append(t)
        doc.build(story)
        buffer.seek(0)
        return buffer

    # =========================================================================
    # 3-SHEET MULTISHEET EXPORT
    # =========================================================================

    @classmethod
    def export_to_excel_multisheet(cls, records, columns=None):
        """
        Creates a 3-sheet Excel workbook:
          Sheet 1 — Daily Detail   : full attendance records
          Sheet 2 — Monthly Summary: per-employee aggregated statistics
          Sheet 3 — Manual Review  : only Needs Manual Review rows
        """
        wb = openpyxl.Workbook()

        # ── Sheet 1: Daily Detail ─────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Daily Detail"
        cls._write_daily_detail_sheet(ws1, records, columns)

        # ── Sheet 2: Monthly Summary ──────────────────────────────────────────
        ws2 = wb.create_sheet("Monthly Summary")
        cls._write_monthly_summary_sheet(ws2, records)

        # ── Sheet 3: Manual Review ────────────────────────────────────────────
        ws3 = wb.create_sheet("Manual Review")
        nmr_records = [r for r in records if "manual review" in str(cls._get_val(r, "status", "")).lower()]
        cls._write_daily_detail_sheet(ws3, nmr_records, columns)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @classmethod
    def _write_daily_detail_sheet(cls, ws, records, columns=None):
        """Write Daily Detail rows to a worksheet with professional styling."""
        # (header_label, record_key, col_width, alignment)
        DETAIL_COLS = [
            ("NO.",            "_rownum",         5,   "center"),
            ("Emp ID",         "employee_id",     12,  "center"),
            ("Employee Name",  "employee_name",   26,  "left"),
            ("Department",     "department",      18,  "left"),
            ("Gender",         "gender",          9,   "center"),
            ("Day",            "weekday",         11,  "center"),
            ("Date",           "attendance_date", 14,  "center"),
            ("Shift",          "shift",           10,  "center"),
            ("Login",          "first_check_in",  11,  "center"),
            ("Logout Date",    "logout_date",     14,  "center"),
            ("Logout",         "last_check_out",  11,  "center"),
            ("Working Hours",  "working_hours",   14,  "center"),
            ("Overtime Hours", "overtime_hours",  14,  "center"),
            ("Status",         "status",          26,  "left"),
            ("Remarks",        "remarks",         52,  "left_wrap"),
        ]

        # ── Header row ───────────────────────────────────────────────────────
        for col_i, (hdr, _, width, _align) in enumerate(DETAIL_COLS, 1):
            cell = ws.cell(row=1, column=col_i, value=hdr)
            cell.fill      = cls._HDR_FILL
            cell.font      = cls._HDR_FONT
            cell.alignment = cls._CENTER
            cell.border    = cls._THIN_BORD
            ws.column_dimensions[get_column_letter(col_i)].width = width

        ws.row_dimensions[1].height = 26
        ws.freeze_panes = "A2"

        # ── Status → fill/font mapping ────────────────────────────────────────
        def _row_style(status):
            st = str(status or "").lower()
            if "manual review" in st: return cls._FILL_YELLOW, cls._FONT_YELLOW
            if "half day"      in st: return cls._FILL_ORANGE, cls._FONT_ORANGE
            if "absent"        in st: return cls._FILL_RED,    cls._FONT_RED
            if "overnight"     in st or "c shift" in st: return cls._FILL_PURPLE, cls._FONT_PURPLE
            if "present"       in st: return cls._FILL_GREEN,  cls._FONT_GREEN
            return None, None

        # ── Alternating plain row fill ────────────────────────────────────────
        _ALT_ODD  = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        _ALT_EVEN = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for row_i, rec in enumerate(records, 2):
            status     = cls._get_val(rec, "status", "")
            fill, font = _row_style(status)
            plain_fill = _ALT_ODD if row_i % 2 == 0 else _ALT_EVEN

            for col_i, (_, key, _, align_tag) in enumerate(DETAIL_COLS, 1):
                if key == "_rownum":
                    val = str(row_i - 1)
                else:
                    raw = cls._get_val(rec, key, "")
                    val = str(raw) if raw not in (None, "") else "--"

                cell = ws.cell(row=row_i, column=col_i, value=val)
                cell.border = cls._THIN_BORD

                # Alignment
                if align_tag == "left_wrap":
                    cell.alignment = cls._LEFT_WRAP
                elif align_tag == "left":
                    cell.alignment = cls._LEFT
                else:
                    cell.alignment = cls._CENTER

                # Fill & font
                if fill:
                    cell.fill = fill
                    if font:
                        # Keep font for status rows but normal size
                        cell.font = font
                else:
                    cell.fill = plain_fill
                    cell.font = cls._NRM_FONT

            # Row height — taller for wrapped remarks
            remarks_val = cls._get_val(rec, "remarks", "")
            has_long_remarks = len(str(remarks_val)) > 50
            ws.row_dimensions[row_i].height = 32 if has_long_remarks else 20

    @classmethod
    def _write_monthly_summary_sheet(cls, ws, records):
        """Write per-employee Monthly Summary to a worksheet."""
        from collections import defaultdict

        # Aggregate per employee
        emp_data = defaultdict(lambda: {
            "name": "", "dept": "", "gender": "",
            "full_day": 0, "half_day": 0, "absent": 0,
            "nmr": 0, "total_hours": 0.0, "overtime_hours": 0.0,
            "total_days": 0
        })

        for rec in records:
            eid    = str(cls._get_val(rec, "employee_id", ""))
            if not eid or eid in ("--", "None"): continue
            d      = emp_data[eid]
            d["name"]   = cls._get_val(rec, "employee_name", "")
            d["dept"]   = cls._get_val(rec, "department", "")
            d["gender"] = cls._get_val(rec, "gender", "")
            status = str(cls._get_val(rec, "status", "")).lower()
            d["total_days"] += 1
            if "manual review" in status:   d["nmr"]       += 1
            elif "half day"    in status:   d["half_day"]  += 1
            elif "absent"      in status:   d["absent"]    += 1
            elif "present"     in status:   d["full_day"]  += 1
            try:
                d["total_hours"]    += float(cls._get_val(rec, "working_hours_decimal", 0) or 0)
                d["overtime_hours"] += float(cls._get_val(rec, "overtime_hours_decimal", 0) or 0)
            except (TypeError, ValueError):
                pass

        # (header, width, alignment)
        SUMMARY_COLS = [
            ("Emp ID",        11, "center"), ("Employee Name",  28, "left"),
            ("Department",    20, "left"),   ("Gender",         10, "center"),
            ("Total Days",    12, "center"), ("Full Day",       12, "center"),
            ("Half Day",      12, "center"), ("Absent",         10, "center"),
            ("NMR",           10, "center"), ("Total Hrs",      13, "center"),
            ("Overtime Hrs",  14, "center"),
        ]

        # Header
        for col_i, (hdr, width, _) in enumerate(SUMMARY_COLS, 1):
            cell = ws.cell(row=1, column=col_i, value=hdr)
            cell.fill      = cls._SUM_FILL
            cell.font      = cls._SUM_FONT
            cell.alignment = cls._CENTER
            cell.border    = cls._THIN_BORD
            ws.column_dimensions[get_column_letter(col_i)].width = width

        ws.row_dimensions[1].height = 26
        ws.freeze_panes = "A2"

        NMR_FILL = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
        NMR_FONT = Font(name="Calibri", size=10, color="713F12", bold=True)
        ABT_FILL = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid")
        ABT_FONT = Font(name="Calibri", size=10, color="881337", bold=True)

        for row_i, (eid, d) in enumerate(sorted(emp_data.items()), 2):
            total_h = d["total_hours"]
            ot_h    = d["overtime_hours"]
            row_vals = [
                eid, d["name"], d["dept"], d["gender"],
                d["total_days"], d["full_day"], d["half_day"],
                d["absent"], d["nmr"],
                f"{int(total_h)}:{int((total_h % 1)*60):02d}",
                f"{int(ot_h)}:{int((ot_h % 1)*60):02d}",
            ]
            zebra = cls._ZEBRA_ODD if row_i % 2 == 0 else cls._ZEBRA_EVEN
            for col_i, (val, (_, _, align_tag)) in enumerate(zip(row_vals, SUMMARY_COLS), 1):
                cell = ws.cell(row=row_i, column=col_i, value=val)
                cell.border    = cls._THIN_BORD
                cell.alignment = cls._LEFT if align_tag == "left" else cls._CENTER
                cell.fill = zebra
                cell.font = cls._NRM_FONT
                # Highlight NMR count
                if col_i == 9 and d["nmr"] > 0:
                    cell.fill = NMR_FILL
                    cell.font = NMR_FONT
                # Highlight Absent count
                if col_i == 8 and d["absent"] > 0:
                    cell.fill = ABT_FILL
                    cell.font = ABT_FONT
            ws.row_dimensions[row_i].height = 20
