# -*- coding: utf-8 -*-
import openpyxl
from io import BytesIO
from datetime import date
from services.exporter import AttendanceExporter


def test_excel_export_sheets_and_no_vendor_name():
    """Verify that export_to_excel creates two sheets without Vendor Name."""
    sample_records = [
        {
            "employee_id": "EMP001",
            "employee_name": "Alice Smith",
            "gender": "Female",
            "attendance_date": date(2026, 7, 1),
            "shift": "Shift A",
            "first_check_in": "09:00",
            "last_check_out": "17:00",
            "working_hours": "08:00",
            "working_hours_decimal": 8.0,
            "overtime_hours": "01:00",
            "overtime_hours_decimal": 1.0,
            "status": "Present (Full Day)",
        },
        {
            "employee_id": "EMP001",
            "employee_name": "Alice Smith",
            "gender": "Female",
            "attendance_date": date(2026, 7, 5),
            "shift": "Shift B1",
            "first_check_in": "17:30",
            "last_check_out": "01:30",
            "working_hours": "08:00",
            "working_hours_decimal": 8.0,
            "overtime_hours": "00:00",
            "overtime_hours_decimal": 0.0,
            "status": "Present (Full Day)",
        },
        {
            "employee_id": "EMP002",
            "employee_name": "Bob Jones",
            "gender": "Male",
            "attendance_date": date(2026, 7, 10),
            "shift": "Unknown",
            "first_check_in": None,
            "last_check_out": None,
            "working_hours": "00:00",
            "working_hours_decimal": 0.0,
            "overtime_hours": "00:00",
            "overtime_hours_decimal": 0.0,
            "status": "Absent",
        },
    ]

    excel_buf = AttendanceExporter.export_to_excel(sample_records)
    wb = openpyxl.load_workbook(BytesIO(excel_buf.getvalue()))

    # 1. Verify sheet names
    assert wb.sheetnames == ["Daily Detail", "Monthly Summary"]

    # 2. Check Daily Detail headers (No Vendor Name)
    ws_detail = wb["Daily Detail"]
    detail_headers = [ws_detail.cell(row=1, column=col).value for col in range(1, ws_detail.max_column + 1)]
    assert "Vendor Name" not in detail_headers
    assert "VENDOR_MAP" not in detail_headers

    # 3. Check Monthly Summary headers (Exact 18 columns, No Vendor Name)
    ws_summary = wb["Monthly Summary"]
    summary_headers = [ws_summary.cell(row=1, column=col).value for col in range(1, ws_summary.max_column + 1)]

    expected_summary_headers = [
        "Employee ID",
        "First Name",
        "Gender",
        "Total Days",
        "Present Days",
        "Absent Days",
        "Half Days",
        "LOP Days",
        "Shift A Count",
        "General Count",
        "Shift B Count",
        "Shift B1 Count",
        "Shift C Count",
        "Missing Punch-Out Count",
        "Missing Punch-In Count",
        "Needs Manual Review Count",
        "Total Working Hours",
        "Total Overtime Hours",
    ]
    assert summary_headers == expected_summary_headers
    assert "Vendor Name" not in summary_headers

    # 4. Check EMP001 monthly summary row metrics
    # EMP001 dates span 2026-07-01 to 2026-07-05 -> Total Days = 5
    emp1_row = [ws_summary.cell(row=2, column=col).value for col in range(1, ws_summary.max_column + 1)]
    assert str(emp1_row[0]) == "EMP001"
    assert str(emp1_row[1]) == "Alice Smith"
    assert str(emp1_row[2]) == "Female"
    assert str(emp1_row[3]) == "5"  # Total Days (span 07-01 to 07-05)
    assert str(emp1_row[4]) == "2"  # Present Days
    assert str(emp1_row[5]) == "0"  # Absent Days
    assert str(emp1_row[8]) == "1"  # Shift A Count
    assert str(emp1_row[11]) == "1"  # Shift B1 Count
    assert str(emp1_row[16]) == "16:00"  # Total Working Hours
    assert str(emp1_row[17]) == "01:00"  # Total Overtime Hours


def test_filtered_records_export_produces_filtered_summary():
    """Verify that exporting filtered records produces a summary reflecting only those records."""
    sample_records = [
        {
            "employee_id": "EMP002",
            "employee_name": "Bob Jones",
            "gender": "Male",
            "attendance_date": date(2026, 7, 10),
            "shift": "General",
            "first_check_in": "09:00",
            "last_check_out": "17:00",
            "working_hours": "08:00",
            "working_hours_decimal": 8.0,
            "status": "Present (Full Day)",
        }
    ]

    excel_buf = AttendanceExporter.export_to_excel(sample_records)
    wb = openpyxl.load_workbook(BytesIO(excel_buf.getvalue()))

    ws_summary = wb["Monthly Summary"]
    rows = list(ws_summary.iter_rows(values_only=True))
    assert len(rows) == 2  # Header + 1 filtered employee row
    assert str(rows[1][0]) == "EMP002"
    assert str(rows[1][1]) == "Bob Jones"
    assert str(rows[1][3]) == "1"  # Total Days
    assert str(rows[1][4]) == "1"  # Present Days
    assert str(rows[1][9]) == "1"  # General Count
